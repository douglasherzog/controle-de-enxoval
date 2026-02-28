import csv
import io
import os
from datetime import UTC, datetime, timedelta
from io import StringIO

import qrcode
from flask import Blueprint, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table
from sqlalchemy import func, or_, text

from .models import (
    Colaborador,
    Configuracao,
    EnxovalItem,
    Movimentacao,
    Revisao,
    Setor,
    Tamanho,
    TipoPeca,
    db,
)

STATUS_OPTIONS = [
    "estoque",
    "entregue",
    "em_uso",
    "em_lavagem",
    "disponivel",
    "extraviado",
]
ALERT_STATUS = {"entregue", "em_uso"}
ALERT_ATENCAO_DIAS = 2
ALERT_CRITICO_DIAS = 4
STATUS_COLORS = {
    "estoque": "#2d6a4f",
    "entregue": "#d97706",
    "em_uso": "#2563eb",
    "em_lavagem": "#0ea5e9",
    "disponivel": "#14b8a6",
    "extraviado": "#dc2626",
}


main_bp = Blueprint("main", __name__)
APP_START = datetime.now(UTC)


def _format_uptime(delta: timedelta) -> str:
    total_seconds = int(delta.total_seconds())
    dias, resto = divmod(total_seconds, 86400)
    horas, resto = divmod(resto, 3600)
    minutos, _ = divmod(resto, 60)
    if dias:
        return f"{dias}d {horas}h {minutos}m"
    if horas:
        return f"{horas}h {minutos}m"
    return f"{minutos}m"


def _obter_configuracao() -> Configuracao:
    config = Configuracao.query.order_by(Configuracao.id.asc()).first()
    if not config:
        config = Configuracao(periodicidade_revisao_dias=7)
        db.session.add(config)
        db.session.commit()
    return config


def _criar_item(
    *,
    nome: str,
    codigo: str,
    tag_rfid: str | None,
    tamanho: str,
    tamanho_customizado: str | None,
    descricao: str | None,
    colaborador: str | None,
    setor: str | None,
    status: str,
    observacao: str,
) -> None:
    item = EnxovalItem(
        nome=nome,
        codigo=codigo,
        tag_rfid=tag_rfid,
        tamanho=tamanho,
        tamanho_customizado=tamanho_customizado,
        descricao=descricao,
        colaborador=colaborador,
        setor=setor,
        status=status,
    )
    db.session.add(item)
    db.session.add(
        Movimentacao(
            item=item,
            status=status,
            colaborador=colaborador,
            setor=setor,
            observacao=observacao,
        )
    )


@main_bp.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        codigo = (request.form.get("codigo") or "").strip().upper()
        tag_rfid = (request.form.get("tag_rfid") or "").strip().upper() or None
        tamanho = (request.form.get("tamanho") or "").strip().upper()
        tamanho_customizado = (request.form.get("tamanho_customizado") or "").strip() or None
        descricao = (request.form.get("descricao") or "").strip() or None
        colaborador = (request.form.get("colaborador") or "").strip() or None
        setor = (request.form.get("setor") or "").strip() or None
        novo_setor_nome = (request.form.get("novo_setor_nome") or "").strip()
        novo_tipo_nome = (request.form.get("novo_tipo_nome") or "").strip()
        novo_colaborador_nome = (request.form.get("novo_colaborador_nome") or "").strip()

        # Handle new tipo peca creation
        if nome == "__novo__" and novo_tipo_nome:
            existente = TipoPeca.query.filter(
                func.lower(TipoPeca.nome) == novo_tipo_nome.lower()
            ).one_or_none()
            if existente:
                if not existente.ativo:
                    existente.ativo = True
                    db.session.add(existente)
                    db.session.commit()
                nome = existente.nome
            else:
                novo_tipo = TipoPeca(nome=novo_tipo_nome)
                db.session.add(novo_tipo)
                db.session.commit()
                nome = novo_tipo.nome

        # Handle new colaborador creation
        if colaborador == "__novo__" and novo_colaborador_nome:
            novo_colaborador_telefone = (
                request.form.get("novo_colaborador_telefone") or ""
            ).strip() or None
            existente = Colaborador.query.filter(
                func.lower(Colaborador.nome) == novo_colaborador_nome.lower()
            ).one_or_none()
            if existente:
                if not existente.ativo:
                    existente.ativo = True
                    existente.telefone = novo_colaborador_telefone or existente.telefone
                    db.session.add(existente)
                    db.session.commit()
                colaborador = existente.nome
            else:
                novo_colaborador = Colaborador(
                    nome=novo_colaborador_nome,
                    telefone=novo_colaborador_telefone,
                )
                db.session.add(novo_colaborador)
                db.session.commit()
                colaborador = novo_colaborador.nome

        # Handle new sector creation
        if setor == "__novo__" and novo_setor_nome:
            existente = Setor.query.filter(
                func.lower(Setor.nome) == novo_setor_nome.lower()
            ).one_or_none()
            if existente:
                if not existente.ativo:
                    existente.ativo = True
                    db.session.add(existente)
                    db.session.commit()
                setor = existente.nome
            else:
                novo_setor = Setor(nome=novo_setor_nome)
                db.session.add(novo_setor)
                db.session.commit()
                setor = novo_setor.nome

        # Handle new tamanho creation
        novo_tamanho_nome = (request.form.get("novo_tamanho_nome") or "").strip()
        if tamanho == "__novo__" and novo_tamanho_nome:
            existente = Tamanho.query.filter(
                func.lower(Tamanho.nome) == novo_tamanho_nome.lower()
            ).one_or_none()
            if existente:
                if not existente.ativo:
                    existente.ativo = True
                    db.session.add(existente)
                    db.session.commit()
                tamanho = existente.nome
            else:
                novo_tamanho = Tamanho(nome=novo_tamanho_nome.upper())
                db.session.add(novo_tamanho)
                db.session.commit()
                tamanho = novo_tamanho.nome

        if nome and codigo and tamanho:
            _criar_item(
                nome=nome,
                codigo=codigo,
                tag_rfid=tag_rfid,
                tamanho=tamanho,
                tamanho_customizado=tamanho_customizado,
                descricao=descricao,
                colaborador=colaborador,
                setor=setor,
                status="estoque",
                observacao="Cadastro inicial",
            )
            db.session.commit()
        return redirect(url_for("main.index"))

    busca = (request.args.get("busca") or "").strip()
    filtro_status = (request.args.get("status") or "").strip()
    filtro_setor = (request.args.get("setor") or "").strip()
    filtro_colaborador = (request.args.get("colaborador") or "").strip()
    apenas_ativos = request.args.get("ativos", "1") == "1"
    try:
        pagina = max(int(request.args.get("pagina", "1")), 1)
    except ValueError:
        pagina = 1
    try:
        por_pagina = max(int(request.args.get("por_pagina", "25")), 5)
    except ValueError:
        por_pagina = 25

    itens_query = EnxovalItem.query
    if apenas_ativos:
        itens_query = itens_query.filter(EnxovalItem.ativo.is_(True))
    if busca:
        termo = f"%{busca}%"
        itens_query = itens_query.filter(
            (EnxovalItem.nome.ilike(termo)) | (EnxovalItem.codigo.ilike(termo))
        )
    if filtro_status:
        itens_query = itens_query.filter(EnxovalItem.status == filtro_status)
    if filtro_setor:
        if filtro_setor == "__sem__":
            itens_query = itens_query.filter(
                or_(EnxovalItem.setor.is_(None), EnxovalItem.setor == "")
            )
        else:
            itens_query = itens_query.filter(EnxovalItem.setor == filtro_setor)
    if filtro_colaborador:
        if filtro_colaborador == "__sem__":
            itens_query = itens_query.filter(
                or_(EnxovalItem.colaborador.is_(None), EnxovalItem.colaborador == "")
            )
        else:
            itens_query = itens_query.filter(EnxovalItem.colaborador == filtro_colaborador)

    total_itens = itens_query.count()
    total_paginas = max((total_itens + por_pagina - 1) // por_pagina, 1)
    if pagina > total_paginas:
        pagina = total_paginas
    offset = (pagina - 1) * por_pagina
    itens = (
        itens_query.order_by(EnxovalItem.id.desc())
        .offset(offset)
        .limit(por_pagina)
        .all()
    )
    tipos_peca_ativos = TipoPeca.query.filter_by(
        ativo=True
    ).order_by(TipoPeca.nome.asc()).all()
    colaboradores_ativos = (
        Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome.asc()).all()
    )
    setores_ativos = Setor.query.filter_by(ativo=True).order_by(Setor.nome.asc()).all()
    tamanhos_ativos = Tamanho.query.filter_by(ativo=True).order_by(Tamanho.nome.asc()).all()
    setores = Setor.query.order_by(Setor.nome.asc()).all()
    config = _obter_configuracao()
    limite_revisao = datetime.now(UTC) - timedelta(days=config.periodicidade_revisao_dias)
    ultima_revisao_sub = (
        db.session.query(
            Revisao.item_id,
            func.max(Revisao.created_at).label("ultima_revisao"),
        )
        .group_by(Revisao.item_id)
        .subquery()
    )
    pendentes_revisao = (
        db.session.query(func.count(EnxovalItem.id))
        .outerjoin(ultima_revisao_sub, EnxovalItem.id == ultima_revisao_sub.c.item_id)
        .filter(EnxovalItem.ativo.is_(True))
        .filter(
            or_(
                ultima_revisao_sub.c.ultima_revisao.is_(None),
                ultima_revisao_sub.c.ultima_revisao < limite_revisao,
            )
        )
        .scalar()
        or 0
    )
    return render_template(
        "index.html",
        itens=itens,
        status_options=STATUS_OPTIONS,
        tipos_peca_ativos=tipos_peca_ativos,
        colaboradores_ativos=colaboradores_ativos,
        setores=setores,
        setores_ativos=setores_ativos,
        tamanhos_ativos=tamanhos_ativos,
        busca=busca,
        filtro_status=filtro_status,
        filtro_setor=filtro_setor,
        filtro_colaborador=filtro_colaborador,
        apenas_ativos=apenas_ativos,
        pagina=pagina,
        por_pagina=por_pagina,
        total_itens=total_itens,
        total_paginas=total_paginas,
        offset=offset,
        pendentes_revisao=pendentes_revisao,
        periodicidade_revisao=config.periodicidade_revisao_dias,
    )


def _montar_dashboard_context() -> dict:
    status_counts = dict(
        db.session.query(EnxovalItem.status, func.count(EnxovalItem.id))
        .group_by(EnxovalItem.status)
        .all()
    )
    pendentes_status = {"entregue", "em_uso", "em_lavagem"}
    pendentes = sum(status_counts.get(status, 0) for status in pendentes_status)
    extraviados = status_counts.get("extraviado", 0)
    total_ativos = (
        db.session.query(func.count(EnxovalItem.id))
        .filter(EnxovalItem.ativo.is_(True))
        .scalar()
        or 0
    )

    por_tipo = (
        db.session.query(EnxovalItem.nome, func.count(EnxovalItem.id))
        .filter(EnxovalItem.ativo.is_(True))
        .group_by(EnxovalItem.nome)
        .order_by(EnxovalItem.nome.asc())
        .all()
    )

    setor_expr = func.coalesce(func.nullif(EnxovalItem.setor, ""), "Sem setor").label("setor")
    por_setor = (
        db.session.query(setor_expr, func.count(EnxovalItem.id))
        .filter(EnxovalItem.ativo.is_(True))
        .group_by(setor_expr)
        .order_by(setor_expr.asc())
        .all()
    )

    ultima_mov_subquery = (
        db.session.query(
            Movimentacao.item_id,
            func.max(Movimentacao.created_at).label("ultima_mov"),
        )
        .group_by(Movimentacao.item_id)
        .subquery()
    )
    itens_alerta = (
        db.session.query(EnxovalItem, ultima_mov_subquery.c.ultima_mov)
        .join(ultima_mov_subquery, EnxovalItem.id == ultima_mov_subquery.c.item_id)
        .filter(EnxovalItem.ativo.is_(True), EnxovalItem.status.in_(ALERT_STATUS))
        .all()
    )
    agora = datetime.now(UTC)
    alerta_total = {"atencao": 0, "critico": 0}
    alerta_por_setor: dict[str, dict[str, int]] = {}
    alerta_por_colaborador: dict[str, dict[str, int]] = {}

    for item, ultima_mov in itens_alerta:
        if not ultima_mov:
            continue
        if ultima_mov.tzinfo is None:
            ultima_mov = ultima_mov.replace(tzinfo=UTC)
        dias = (agora - ultima_mov).days
        nivel = None
        if dias > ALERT_CRITICO_DIAS:
            nivel = "critico"
        elif dias > ALERT_ATENCAO_DIAS:
            nivel = "atencao"
        if not nivel:
            continue

        alerta_total[nivel] += 1
        setor = item.setor or "Sem setor"
        colaborador = item.colaborador or "Sem colaborador"
        alerta_por_setor.setdefault(setor, {"atencao": 0, "critico": 0})[nivel] += 1
        alerta_por_colaborador.setdefault(colaborador, {"atencao": 0, "critico": 0})[nivel] += 1

    def _ordenar_alertas(dados: dict[str, dict[str, int]]) -> list[tuple[str, int, int, int]]:
        resultado = []
        for chave, valores in dados.items():
            atencao = valores.get("atencao", 0)
            critico = valores.get("critico", 0)
            resultado.append((chave, atencao, critico, atencao + critico))
        return sorted(resultado, key=lambda item: item[3], reverse=True)

    alertas_setor = _ordenar_alertas(alerta_por_setor)
    alertas_colaborador = _ordenar_alertas(alerta_por_colaborador)

    status_chart = []
    total_status = sum(status_counts.values()) or 1
    acumulado = 0.0
    segmentos = []
    for status in STATUS_OPTIONS:
        quantidade = status_counts.get(status, 0)
        if quantidade == 0:
            continue
        percentual = (quantidade / total_status) * 100
        status_chart.append(
            {
                "status": status,
                "quantidade": quantidade,
                "percentual": percentual,
                "color": STATUS_COLORS.get(status, "#94a3b8"),
            }
        )
        inicio = acumulado
        fim = acumulado + percentual
        segmentos.append(
            f"{STATUS_COLORS.get(status, '#94a3b8')} {inicio:.2f}% {fim:.2f}%"
        )
        acumulado = fim
    status_conic = (
        f"conic-gradient({', '.join(segmentos)})"
        if segmentos
        else "conic-gradient(#e2e8f0 0% 100%)"
    )
    max_tipo = max((total for _, total in por_tipo), default=1)
    max_setor = max((total for _, total in por_setor), default=1)

    return {
        "status_counts": status_counts,
        "pendentes": pendentes,
        "extraviados": extraviados,
        "total_ativos": total_ativos,
        "por_tipo": por_tipo,
        "por_setor": por_setor,
        "alerta_total": alerta_total,
        "alertas_setor": alertas_setor,
        "alertas_colaborador": alertas_colaborador,
        "status_chart": status_chart,
        "status_conic": status_conic,
        "max_tipo": max_tipo,
        "max_setor": max_setor,
    }


@main_bp.route("/dashboard")
def dashboard():
    """Página com gráficos e indicadores do enxoval."""
    contexto = _montar_dashboard_context()
    return render_template("dashboard.html", **contexto)


@main_bp.route("/status")
def status():
    """Página simples de status do sistema."""
    version = os.getenv("APP_VERSION", "dev")
    uptime = _format_uptime(datetime.now(UTC) - APP_START)
    db_ok = True
    db_error = None
    try:
        db.session.execute(text("SELECT 1"))
    except Exception as exc:  # noqa: BLE001
        db_ok = False
        db_error = str(exc)

    return render_template(
        "status.html",
        version=version,
        uptime=uptime,
        db_ok=db_ok,
        db_error=db_error,
    )


@main_bp.route("/revisao", methods=["GET", "POST"])
def revisao():
    """Tela de revisão rápida do enxoval."""
    config = _obter_configuracao()
    filtro_setor = (request.args.get("setor") or "").strip()
    filtro_colaborador = (request.args.get("colaborador") or "").strip()

    if request.method == "POST":
        acao = request.form.get("acao")
        next_url = request.form.get("next") or url_for(
            "main.revisao",
            setor=filtro_setor,
            colaborador=filtro_colaborador,
        )
        if acao == "config":
            try:
                periodicidade = int(request.form.get("periodicidade") or 0)
            except ValueError:
                periodicidade = config.periodicidade_revisao_dias
            if periodicidade <= 0:
                periodicidade = config.periodicidade_revisao_dias
            config.periodicidade_revisao_dias = periodicidade
            db.session.add(config)
            db.session.commit()
            return redirect(next_url)
        if acao == "conferir":
            item_id = request.form.get("item_id")
            conferente = (request.form.get("conferente") or "").strip()
            item = db.session.get(EnxovalItem, item_id)
            if item and conferente:
                revisao = Revisao(
                    item_id=item.id,
                    conferente=conferente,
                    setor=item.setor,
                    colaborador=item.colaborador,
                )
                db.session.add(revisao)
                db.session.commit()
            return redirect(next_url)

    limite = datetime.now(UTC) - timedelta(days=config.periodicidade_revisao_dias)
    ultima_revisao_sub = (
        db.session.query(
            Revisao.item_id,
            func.max(Revisao.created_at).label("ultima_revisao"),
        )
        .group_by(Revisao.item_id)
        .subquery()
    )

    itens_query = (
        db.session.query(EnxovalItem, ultima_revisao_sub.c.ultima_revisao)
        .outerjoin(ultima_revisao_sub, EnxovalItem.id == ultima_revisao_sub.c.item_id)
        .filter(EnxovalItem.ativo.is_(True))
        .filter(
            or_(
                ultima_revisao_sub.c.ultima_revisao.is_(None),
                ultima_revisao_sub.c.ultima_revisao < limite,
            )
        )
    )

    if filtro_setor:
        if filtro_setor == "__sem__":
            itens_query = itens_query.filter(
                or_(EnxovalItem.setor.is_(None), EnxovalItem.setor == "")
            )
        else:
            itens_query = itens_query.filter(EnxovalItem.setor == filtro_setor)
    if filtro_colaborador:
        if filtro_colaborador == "__sem__":
            itens_query = itens_query.filter(
                or_(EnxovalItem.colaborador.is_(None), EnxovalItem.colaborador == "")
            )
        else:
            itens_query = itens_query.filter(EnxovalItem.colaborador == filtro_colaborador)

    itens = itens_query.order_by(EnxovalItem.id.desc()).all()
    setores_ativos = Setor.query.filter_by(ativo=True).order_by(Setor.nome.asc()).all()
    colaboradores_ativos = (
        Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome.asc()).all()
    )

    return render_template(
        "revisao.html",
        itens=itens,
        config=config,
        filtro_setor=filtro_setor,
        filtro_colaborador=filtro_colaborador,
        setores_ativos=setores_ativos,
        colaboradores_ativos=colaboradores_ativos,
    )


@main_bp.route("/revisao/scan", methods=["GET", "POST"])
def revisao_scan():
    """Leitura por QR Code para revisão rápida."""
    if request.method == "GET":
        return render_template("revisao_scan.html")

    dados = request.get_json() or {}
    raw = (dados.get("raw") or "").strip()
    codigo = (dados.get("codigo") or "").strip().upper()
    conferente = (dados.get("conferente") or "").strip()

    if not codigo and raw:
        partes = raw.split("|")
        for parte in partes:
            if parte.upper().startswith("CODIGO:"):
                codigo = parte.split(":", 1)[1].strip().upper()
                break
        if not codigo:
            codigo = raw.strip().upper()

    if not codigo:
        return jsonify({"sucesso": False, "mensagem": "Código não identificado."}), 400
    if not conferente:
        return jsonify({"sucesso": False, "mensagem": "Informe o conferente."}), 400

    item = EnxovalItem.query.filter_by(codigo=codigo).first()
    if not item:
        return jsonify(
            {"sucesso": False, "mensagem": f"Peça {codigo} não encontrada."}
        ), 404

    revisao = Revisao(
        item_id=item.id,
        conferente=conferente,
        setor=item.setor,
        colaborador=item.colaborador,
    )
    db.session.add(revisao)
    db.session.commit()

    return jsonify(
        {
            "sucesso": True,
            "mensagem": f"Peça {item.codigo} conferida.",
            "item": {
                "codigo": item.codigo,
                "nome": item.nome,
                "tamanho": item.tamanho,
                "setor": item.setor,
                "colaborador": item.colaborador,
            },
        }
    )


@main_bp.route("/revisoes/relatorio")
def relatorio_revisoes():
    """Relatório simples de revisões por período."""
    try:
        dias = int(request.args.get("dias", "7"))
    except ValueError:
        dias = 7
    if dias <= 0:
        dias = 7

    inicio = datetime.now(UTC) - timedelta(days=dias)
    revisoes = (
        Revisao.query.filter(Revisao.created_at >= inicio)
        .order_by(Revisao.created_at.desc())
        .all()
    )
    total = len(revisoes)
    por_conferente = (
        db.session.query(Revisao.conferente, func.count(Revisao.id))
        .filter(Revisao.created_at >= inicio)
        .group_by(Revisao.conferente)
        .order_by(func.count(Revisao.id).desc())
        .all()
    )

    return render_template(
        "relatorio_revisoes.html",
        revisoes=revisoes,
        total=total,
        por_conferente=por_conferente,
        dias=dias,
        inicio=inicio,
    )


@main_bp.route("/item/<int:item_id>")
def item_detalhe(item_id: int):
    item = db.session.get(EnxovalItem, item_id)
    if not item:
        return redirect(url_for("main.index"))

    movimentacoes = (
        Movimentacao.query.filter_by(item_id=item_id)
        .order_by(Movimentacao.created_at.desc())
        .all()
    )
    ultima_revisao = (
        Revisao.query.filter_by(item_id=item_id)
        .order_by(Revisao.created_at.desc())
        .first()
    )
    revisoes_recentes = (
        Revisao.query.filter_by(item_id=item_id)
        .order_by(Revisao.created_at.desc())
        .limit(5)
        .all()
    )
    return render_template(
        "item.html",
        item=item,
        movimentacoes=movimentacoes,
        ultima_revisao=ultima_revisao,
        revisoes_recentes=revisoes_recentes,
    )


@main_bp.route("/item/<int:item_id>/editar", methods=["GET", "POST"])
def editar_item(item_id: int):
    """Edita uma peça existente."""
    item = db.session.get(EnxovalItem, item_id)
    if not item:
        return redirect(url_for("main.index"))

    if request.method == "POST":
        item.nome = (request.form.get("nome") or item.nome).strip()
        item.codigo = (request.form.get("codigo") or item.codigo).strip().upper()
        item.tamanho = (request.form.get("tamanho") or item.tamanho).strip().upper()
        item.tamanho_customizado = (request.form.get("tamanho_customizado") or "").strip() or None
        item.tag_rfid = (request.form.get("tag_rfid") or "").strip().upper() or None
        item.descricao = (request.form.get("descricao") or "").strip() or None
        item.colaborador = (request.form.get("colaborador") or "").strip() or None
        item.setor = (request.form.get("setor") or "").strip() or None

        db.session.add(item)
        db.session.commit()
        return redirect(url_for("main.index"))

    # GET - mostrar formulário de edição
    setores_ativos = Setor.query.filter_by(ativo=True).order_by(Setor.nome.asc()).all()
    colaboradores_ativos = (
        Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome.asc()).all()
    )
    tipos_peca_ativos = (
        TipoPeca.query.filter_by(ativo=True).order_by(TipoPeca.nome.asc()).all()
    )
    tamanhos_ativos = Tamanho.query.filter_by(ativo=True).order_by(Tamanho.nome.asc()).all()

    return render_template(
        "editar_item.html",
        item=item,
        setores_ativos=setores_ativos,
        colaboradores_ativos=colaboradores_ativos,
        tipos_peca_ativos=tipos_peca_ativos,
        tamanhos_ativos=tamanhos_ativos,
    )


@main_bp.route("/colaboradores/<int:colaborador_id>/editar", methods=["GET", "POST"])
def editar_colaborador(colaborador_id: int):
    """Edita um colaborador existente."""
    colaborador = db.session.get(Colaborador, colaborador_id)
    if not colaborador:
        return redirect(url_for("main.index"))

    if request.method == "POST":
        colaborador.nome = (request.form.get("nome") or colaborador.nome).strip()
        colaborador.telefone = (request.form.get("telefone") or "").strip() or None

        db.session.add(colaborador)
        db.session.commit()
        return redirect(url_for("main.gerenciar_colaboradores"))

    return render_template("editar_colaborador.html", colaborador=colaborador)


@main_bp.route("/setores/<int:setor_id>/editar", methods=["GET", "POST"])
def editar_setor(setor_id: int):
    """Edita um setor existente."""
    setor = db.session.get(Setor, setor_id)
    if not setor:
        return redirect(url_for("main.index"))

    if request.method == "POST":
        setor.nome = (request.form.get("nome") or setor.nome).strip()

        db.session.add(setor)
        db.session.commit()
        return redirect(url_for("main.gerenciar_setores"))

    return render_template("editar_setor.html", setor=setor)


@main_bp.route("/tipos-peca/<int:tipo_id>/editar", methods=["GET", "POST"])
def editar_tipo_peca(tipo_id: int):
    """Edita um tipo de peça existente."""
    tipo = db.session.get(TipoPeca, tipo_id)
    if not tipo:
        return redirect(url_for("main.index"))

    if request.method == "POST":
        tipo.nome = (request.form.get("nome") or tipo.nome).strip()

        db.session.add(tipo)
        db.session.commit()
        return redirect(url_for("main.index"))

    return render_template("editar_tipo_peca.html", tipo=tipo)


@main_bp.route("/importar", methods=["POST"])
def importar_csv():
    """Importa dados de CSV."""
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return redirect(url_for("main.index"))

    conteudo = arquivo.stream.read().decode("utf-8-sig")
    leitor = csv.DictReader(StringIO(conteudo))
    for linha in leitor:
        nome = (linha.get("nome") or "").strip()
        codigo = (linha.get("codigo") or "").strip().upper()
        tag_rfid = (linha.get("tag_rfid") or "").strip().upper() or None
        tamanho = (linha.get("tamanho") or "").strip().upper()
        tamanho_customizado = (linha.get("tamanho_customizado") or "").strip() or None
        descricao = (linha.get("descricao") or "").strip() or None
        colaborador = (linha.get("colaborador") or "").strip() or None
        setor = (linha.get("setor") or "").strip() or None
        status = (linha.get("status") or "").strip().lower() or "estoque"
        observacao = (linha.get("observacao") or "").strip() or "Importacao CSV"

        if not nome or not codigo or not tamanho:
            continue
        if status not in STATUS_OPTIONS:
            status = "estoque"

        _criar_item(
            nome=nome,
            codigo=codigo,
            tag_rfid=tag_rfid,
            tamanho=tamanho,
            tamanho_customizado=tamanho_customizado,
            descricao=descricao,
            colaborador=colaborador,
            setor=setor,
            status=status,
            observacao=observacao,
        )

    db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/inativar/<int:item_id>", methods=["POST"])
def inativar_item(item_id: int):
    item = db.session.get(EnxovalItem, item_id)
    if item:
        item.ativo = False
        db.session.add(item)
        db.session.commit()
    return redirect(request.referrer or url_for("main.index"))


@main_bp.route("/movimentar/<int:item_id>", methods=["POST"])
def movimentar_item(item_id: int):
    item = db.session.get(EnxovalItem, item_id)
    if not item:
        return redirect(url_for("main.index"))

    status = (request.form.get("status") or "").strip()
    colaborador = (request.form.get("colaborador") or "").strip() or None
    setor = (request.form.get("setor") or "").strip() or None
    observacao = (request.form.get("observacao") or "").strip() or None

    if status in STATUS_OPTIONS:
        item.status = status
        item.colaborador = colaborador
        item.setor = setor
        db.session.add(item)
        db.session.add(
            Movimentacao(
                item=item,
                status=status,
                colaborador=colaborador,
                setor=setor,
                observacao=observacao,
            )
        )
        db.session.commit()

    return redirect(url_for("main.index"))


@main_bp.route("/setores", methods=["POST"])
def criar_setor():
    nome = (request.form.get("nome") or "").strip()
    if not nome:
        return redirect(url_for("main.index"))

    existente = (
        Setor.query.filter(func.lower(Setor.nome) == nome.lower()).one_or_none()
    )
    if existente:
        if not existente.ativo:
            existente.ativo = True
            db.session.add(existente)
            db.session.commit()
        return redirect(url_for("main.index"))

    setor = Setor(nome=nome)
    db.session.add(setor)
    db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/setores/<int:setor_id>/inativar", methods=["POST"])
def inativar_setor(setor_id: int):
    setor = db.session.get(Setor, setor_id)
    if setor:
        setor.ativo = False
        db.session.add(setor)
        db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/tipos-peca", methods=["POST"])
def criar_tipo_peca():
    nome = (request.form.get("nome") or "").strip()
    if not nome:
        return redirect(url_for("main.index"))

    existente = (
        TipoPeca.query.filter(func.lower(TipoPeca.nome) == nome.lower()).one_or_none()
    )
    if existente:
        if not existente.ativo:
            existente.ativo = True
            db.session.add(existente)
            db.session.commit()
        return redirect(url_for("main.index"))

    tipo_peca = TipoPeca(nome=nome)
    db.session.add(tipo_peca)
    db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/tipos-peca/<int:tipo_id>/inativar", methods=["POST"])
def inativar_tipo_peca(tipo_id: int):
    tipo_peca = db.session.get(TipoPeca, tipo_id)
    if tipo_peca:
        tipo_peca.ativo = False
        db.session.add(tipo_peca)
        db.session.commit()
        return redirect(url_for("main.gerenciar_tipos_peca"))

    return redirect(url_for("main.index"))


@main_bp.route("/colaboradores", methods=["POST"])
def criar_colaborador():
    nome = (request.form.get("nome") or "").strip()
    if not nome:
        return redirect(url_for("main.index"))

    existente = (
        Colaborador.query.filter(func.lower(Colaborador.nome) == nome.lower()).one_or_none()
    )
    if existente:
        if not existente.ativo:
            existente.ativo = True
            db.session.add(existente)
            db.session.commit()
        return redirect(url_for("main.index"))

    colaborador = Colaborador(nome=nome)
    db.session.add(colaborador)
    db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/colaboradores/<int:colaborador_id>/inativar", methods=["POST"])
def inativar_colaborador(colaborador_id: int):
    colaborador = db.session.get(Colaborador, colaborador_id)
    if colaborador:
        colaborador.ativo = False
        db.session.add(colaborador)
        db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/relatorio/<periodo>")
def relatorio_status(periodo: str):
    """Gera relatório de status (diario, semanal, mensal)."""
    from datetime import timedelta

    agora = datetime.now(UTC)

    if periodo == "diario":
        data_inicio = agora - timedelta(days=1)
        titulo = "Relatório Diário"
    elif periodo == "semanal":
        data_inicio = agora - timedelta(weeks=1)
        titulo = "Relatório Semanal"
    elif periodo == "mensal":
        data_inicio = agora - timedelta(days=30)
        titulo = "Relatório Mensal"
    else:
        return redirect(url_for("main.index"))

    # Estatísticas gerais
    status_counts = dict(
        db.session.query(EnxovalItem.status, func.count(EnxovalItem.id))
        .filter(EnxovalItem.ativo.is_(True))
        .group_by(EnxovalItem.status)
        .all()
    )

    # Movimentações no período
    movimentacoes = (
        Movimentacao.query.filter(Movimentacao.created_at >= data_inicio)
        .order_by(Movimentacao.created_at.desc())
        .limit(100)
        .all()
    )

    # Itens com alerta
    ultima_mov_subquery = (
        db.session.query(
            Movimentacao.item_id,
            func.max(Movimentacao.created_at).label("ultima_mov"),
        )
        .group_by(Movimentacao.item_id)
        .subquery()
    )

    itens_alerta = (
        db.session.query(EnxovalItem, ultima_mov_subquery.c.ultima_mov)
        .join(ultima_mov_subquery, EnxovalItem.id == ultima_mov_subquery.c.item_id)
        .filter(
            EnxovalItem.ativo.is_(True),
            EnxovalItem.status.in_(ALERT_STATUS),
        )
        .all()
    )

    alertas = []
    for item, ultima_mov in itens_alerta:
        if not ultima_mov:
            continue
        if ultima_mov.tzinfo is None:
            ultima_mov = ultima_mov.replace(tzinfo=UTC)
        dias = (agora - ultima_mov).days
        if dias > ALERT_CRITICO_DIAS:
            alertas.append((item, dias, "critico"))
        elif dias > ALERT_ATENCAO_DIAS:
            alertas.append((item, dias, "atencao"))

    # Agrupar por tipo e setor
    por_tipo = (
        db.session.query(EnxovalItem.nome, func.count(EnxovalItem.id))
        .filter(EnxovalItem.ativo.is_(True))
        .group_by(EnxovalItem.nome)
        .order_by(func.count(EnxovalItem.id).desc())
        .limit(10)
        .all()
    )

    setor_expr = func.coalesce(func.nullif(EnxovalItem.setor, ""), "Sem setor")
    por_setor = (
        db.session.query(
            setor_expr.label("setor"),
            func.count(EnxovalItem.id),
        )
        .filter(EnxovalItem.ativo.is_(True))
        .group_by(setor_expr)
        .order_by(func.count(EnxovalItem.id).desc())
        .limit(10)
        .all()
    )

    total_ativos = sum(status_counts.values())

    return render_template(
        "relatorio.html",
        titulo=titulo,
        periodo=periodo,
        data_inicio=data_inicio,
        data_fim=agora,
        status_counts=status_counts,
        total_ativos=total_ativos,
        movimentacoes=movimentacoes,
        alertas=alertas,
        por_tipo=por_tipo,
        por_setor=por_setor,
        status_options=STATUS_OPTIONS,
        status_colors=STATUS_COLORS,
    )


def seed_tipos_peca():
    """Pré-cadastra os tipos de peças padrão do frigorífico."""
    tipos_padrao = [
        "Moletons",
        "Calças brancas",
        "Calças NR10",
        "Calças térmicas",
        "Camisas NR10",
        "Camisetas com capuz acoplado",
        "Capuz",
        "Capuz térmico",
        "Jaquetas térmicas",
        "Batas manga curta branca",
        "Batas manga longa branca",
        "Batas manga curta azul",
        "Batas manga longa cinza",
        "Calças azuis",
        "Calças cinzas",
    ]

    for tipo_nome in tipos_padrao:
        existente = TipoPeca.query.filter(
            func.lower(TipoPeca.nome) == tipo_nome.lower()
        ).one_or_none()
        if not existente:
            novo_tipo = TipoPeca(nome=tipo_nome, ativo=True)
            db.session.add(novo_tipo)

    db.session.commit()


@main_bp.route("/qrcode/<int:item_id>")
def gerar_qrcode(item_id: int):
    """Gera QR code para uma peça específica."""
    item = db.session.get(EnxovalItem, item_id)
    if not item:
        return redirect(url_for("main.index"))

    # Dados para o QR code
    dados = f"CODIGO:{item.codigo}|TIPO:{item.nome}|TAMANHO:{item.tamanho}"
    if item.tag_rfid:
        dados += f"|RFID:{item.tag_rfid}"

    # Gerar QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(dados)
    qr.make(fit=True)

    # Criar imagem
    img = qr.make_image(fill_color="black", back_color="white")

    # Salvar em buffer
    img_buffer = io.BytesIO()
    img.save(img_buffer, format="PNG")
    img_buffer.seek(0)

    return send_file(
        img_buffer,
        mimetype="image/png",
        as_download=False,
    )


@main_bp.route("/etiqueta/<int:item_id>")
def etiqueta_item(item_id: int):
    """Mostra página de etiqueta para impressão."""
    item = db.session.get(EnxovalItem, item_id)
    if not item:
        return redirect(url_for("main.index"))

    return render_template("etiqueta.html", item=item)


@main_bp.route("/relatorio/<periodo>/pdf")
def relatorio_pdf(periodo: str):
    """Gera relatório em PDF."""
    from io import BytesIO

    agora = datetime.now(UTC)

    if periodo == "diario":
        data_inicio = agora - timedelta(days=1)
        titulo = "Relatório Diário"
    elif periodo == "semanal":
        data_inicio = agora - timedelta(weeks=1)
        titulo = "Relatório Semanal"
    elif periodo == "mensal":
        data_inicio = agora - timedelta(days=30)
        titulo = "Relatório Mensal"
    else:
        return redirect(url_for("main.index"))

    # Estatísticas
    status_counts = dict(
        db.session.query(EnxovalItem.status, func.count(EnxovalItem.id))
        .filter(EnxovalItem.ativo.is_(True))
        .group_by(EnxovalItem.status)
        .all()
    )
    total_ativos = sum(status_counts.values())

    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    elements.append(Paragraph(f"<b>{titulo}</b>", styles["Title"]))
    elements.append(
        Paragraph(
            f"Período: {data_inicio.strftime('%d/%m/%Y')} a {agora.strftime('%d/%m/%Y')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 20))

    # Tabela de status
    data = [["Status", "Quantidade", "Percentual"]]
    for status in STATUS_OPTIONS:
        count = status_counts.get(status, 0)
        if count > 0:
            percent = f"{(count / total_ativos * 100):.1f}%"
            data.append([status.replace("_", " "), str(count), percent])

    table = Table(data)
    table.setStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 14),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]
    )
    elements.append(table)
    elements.append(Spacer(1, 20))

    # Total
    elements.append(Paragraph(f"<b>Total de peças ativas: {total_ativos}</b>", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/pdf",
        as_download=True,
        download_name=f"relatorio_{periodo}_{agora.strftime('%Y%m%d')}.pdf",
    )


@main_bp.route("/relatorio/<periodo>/excel")
def relatorio_excel(periodo: str):
    """Gera relatório em Excel."""
    from io import BytesIO

    agora = datetime.now(UTC)

    if periodo == "diario":
        data_inicio = agora - timedelta(days=1)
        titulo = "Relatório Diário"
    elif periodo == "semanal":
        data_inicio = agora - timedelta(weeks=1)
        titulo = "Relatório Semanal"
    elif periodo == "mensal":
        data_inicio = agora - timedelta(days=30)
        titulo = "Relatório Mensal"
    else:
        return redirect(url_for("main.index"))

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Cabeçalho
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Período: {data_inicio.strftime('%d/%m/%Y')} a {agora.strftime('%d/%m/%Y')}"

    # Estatísticas
    status_counts = dict(
        db.session.query(EnxovalItem.status, func.count(EnxovalItem.id))
        .filter(EnxovalItem.ativo.is_(True))
        .group_by(EnxovalItem.status)
        .all()
    )
    total_ativos = sum(status_counts.values())

    ws["A4"] = "Status"
    ws["B4"] = "Quantidade"
    ws["C4"] = "Percentual"
    for cell in [ws["A4"], ws["B4"], ws["C4"]]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row = 5
    for status in STATUS_OPTIONS:
        count = status_counts.get(status, 0)
        if count > 0:
            ws[f"A{row}"] = status.replace("_", " ")
            ws[f"B{row}"] = count
            ws[f"C{row}"] = f"{(count / total_ativos * 100):.1f}%"
            row += 1

    ws[f"A{row + 1}"] = "Total"
    ws[f"A{row + 1}"].font = Font(bold=True)
    ws[f"B{row + 1}"] = total_ativos
    ws[f"B{row + 1}"].font = Font(bold=True)

    # Ajustar larguras
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_download=True,
        download_name=f"relatorio_{periodo}_{agora.strftime('%Y%m%d')}.xlsx",
    )


@main_bp.route("/tamanhos", methods=["POST"])
def criar_tamanho():
    """Cria um novo tamanho."""
    nome = (request.form.get("nome") or "").strip().upper()
    if not nome:
        return redirect(url_for("main.index"))

    existente = (
        Tamanho.query.filter(func.lower(Tamanho.nome) == nome.lower()).one_or_none()
    )
    if existente:
        if not existente.ativo:
            existente.ativo = True
            db.session.add(existente)
            db.session.commit()
        return redirect(url_for("main.index"))

    tamanho = Tamanho(nome=nome)
    db.session.add(tamanho)
    db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/tamanhos/<int:tamanho_id>/editar", methods=["GET", "POST"])
def editar_tamanho(tamanho_id: int):
    """Edita um tamanho existente."""
    tamanho = db.session.get(Tamanho, tamanho_id)
    if not tamanho:
        return redirect(url_for("main.index"))

    if request.method == "POST":
        tamanho.nome = (request.form.get("nome") or tamanho.nome).strip().upper()
        db.session.add(tamanho)
        db.session.commit()
        return redirect(url_for("main.gerenciar_tamanhos"))

    return render_template("editar_tamanho.html", tamanho=tamanho)


@main_bp.route("/tamanhos/<int:tamanho_id>/inativar", methods=["POST"])
def inativar_tamanho(tamanho_id: int):
    """Inativa um tamanho."""
    tamanho = db.session.get(Tamanho, tamanho_id)
    if tamanho:
        tamanho.ativo = False
        db.session.add(tamanho)
        db.session.commit()
    return redirect(url_for("main.index"))


@main_bp.route("/item/<int:item_id>/excluir", methods=["POST"])
def excluir_item(item_id: int):
    """Exclui permanentemente uma peça e suas movimentações."""
    item = db.session.get(EnxovalItem, item_id)
    if item:
        # Excluir movimentações primeiro
        Movimentacao.query.filter_by(item_id=item_id).delete()
        # Excluir item
        db.session.delete(item)
        db.session.commit()
    return redirect(request.referrer or url_for("main.index"))


@main_bp.route("/colaboradores/<int:colaborador_id>/excluir", methods=["POST"])
def excluir_colaborador(colaborador_id: int):
    """Exclui permanentemente um colaborador."""
    colaborador = db.session.get(Colaborador, colaborador_id)
    if colaborador:
        db.session.delete(colaborador)
        db.session.commit()
    return redirect(url_for("main.index"))


def seed_tamanhos():
    """Pré-cadastra os tamanhos padrão."""
    tamanhos_padrao = ["P", "M", "G", "GG", "SOB MEDIDA"]

    for nome in tamanhos_padrao:
        existente = Tamanho.query.filter(
            func.lower(Tamanho.nome) == nome.lower()
        ).one_or_none()
        if not existente:
            novo = Tamanho(nome=nome, ativo=True)
            db.session.add(novo)

    db.session.commit()


# Admin management routes
@main_bp.route("/gerenciar")
def gerenciar():
    """Página principal de gerenciamento de cadastros."""
    colaboradores_ativos = Colaborador.query.filter_by(ativo=True).all()
    setores_ativos = Setor.query.filter_by(ativo=True).all()
    tipos_peca_ativos = TipoPeca.query.filter_by(ativo=True).all()
    tamanhos_ativos = Tamanho.query.filter_by(ativo=True).all()
    return render_template(
        "gerenciar.html",
        colaboradores_ativos=colaboradores_ativos,
        setores_ativos=setores_ativos,
        tipos_peca_ativos=tipos_peca_ativos,
        tamanhos_ativos=tamanhos_ativos,
    )


@main_bp.route("/gerenciar/colaboradores")
def gerenciar_colaboradores():
    """Página de gerenciamento de colaboradores."""
    colaboradores = Colaborador.query.order_by(Colaborador.nome.asc()).all()
    return render_template("gerenciar_colaboradores.html", colaboradores=colaboradores)


@main_bp.route("/gerenciar/setores")
def gerenciar_setores():
    """Página de gerenciamento de setores."""
    setores = Setor.query.order_by(Setor.nome.asc()).all()
    return render_template("gerenciar_setores.html", setores=setores)


@main_bp.route("/gerenciar/tipos")
def gerenciar_tipos_peca():
    """Página de gerenciamento de tipos de peça."""
    tipos = TipoPeca.query.order_by(TipoPeca.nome.asc()).all()
    return render_template("gerenciar_tipos.html", tipos=tipos)


@main_bp.route("/gerenciar/tamanhos")
def gerenciar_tamanhos():
    """Página de gerenciamento de tamanhos."""
    tamanhos = Tamanho.query.order_by(Tamanho.nome.asc()).all()
    return render_template("gerenciar_tamanhos.html", tamanhos=tamanhos)
